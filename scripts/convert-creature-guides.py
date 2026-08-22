"""
Converts the compiled taming workbook into DD-S creature info records.

The workbook is the human's compilation step; this is the mechanical
translation into the schema in src/model/creatureInfo.ts. Output is committed
as scripts/data/creature-info.json and read by the official package build, so
it must be deterministic — ids are derived from position, never generated.

Three things the workbook gets wrong, corrected here rather than by hand:

  Mis-associated guides
    Thirty rows carry a verbatim copy of the previous row's guide, so the
    creature above them is what the text actually describes — Tek Rex holding
    Tek Raptor's guide, Rhyniognatha holding Rex Ghost's. A copy is dropped
    rather than repaired: the real guide for those creatures was never
    captured, and inventing one would be worse than having none.

  Citation markers
    The AI capture left 768 of them ("([ARK Wiki][1])"), which mean nothing
    outside the chat they came from.

  Variant duplication
    A variant only keeps a section that genuinely differs from its base
    creature. Everything else is inherited, which is both what an
    administrator expects and how `resolveCreatureInfo` already works.

Run:  python scripts/convert-creature-guides.py scripts/data/creature-info.json
"""

import json
import re
import sys
from collections import Counter, defaultdict

import openpyxl

WORKBOOK = r"C:\Users\joshu\Desktop\official-ark-taming-guides.xlsx"
HERE = __file__.rsplit("\\", 1)[0].rsplit("/", 1)[0]
CATALOG = f"{HERE}/../src/assets/catalog/official-asa.json"
OUT = sys.argv[1] if len(sys.argv) > 1 else f"{HERE}/data/creature-info.json"

SECTIONS = ["acquisition", "spawns", "abilities", "drops", "technical", "notes"]

# Words that qualify a creature rather than name it. Used to tell a variant
# pair ("Ferox" / "Ferox (Large)") from two unrelated creatures that happen to
# share a guide.
QUALIFIER = (
    r"(aberrant|alpha|tek|ghost|corrupted|brute|eerie|vr|enraged|malfunctioned"
    r"|x-|r-|zombie|skeletal|summoned|mega|\(alpha\)|\(beta\)|\(gamma\)"
    r"|\(large\)|elder|prime)"
)

# Name qualifiers that make one creature a variant of another. Order matters:
# the longest match wins, so "Malfunctioned Tek Rex" finds "Rex" rather than
# stopping at a "Tek Rex" that may not exist.
VARIANT_PREFIXES = [
    "Malfunctioned Tek ", "Malfunctioned ", "Aberrant ", "Corrupted ",
    "Skeletal ", "Summoned ", "Ascended ", "Enraged ", "Zombie ", "Brute ",
    "Alpha ", "Eerie ", "Elder ", "Tek ", "VR ", "X-", "R-",
]
VARIANT_SUFFIXES = [" Ghost", " (Alpha)", " (Beta)", " (Gamma)", " (Large)"]

# Citation markers the AI capture left behind.
CITATIONS = [
    # ([ARK Wiki][1]), ([ARK: Survival Evolved Wiki][2])
    re.compile(r"\s*\(\[[^\]]{1,60}\]\[\d+\]\)"),
    # ([ark.wiki.gg](https://ark.wiki.gg/wiki/Thing?utm_source=chatgpt.com))
    re.compile(r"\s*\(\[[^\]]{1,60}\]\([^)]{1,300}\)\)"),
]


def norm(path):
    return path.strip().lower()


def strip_citations(text):
    for pattern in CITATIONS:
        text = pattern.sub("", text)
    # A citation sat before the full stop as often as after it.
    return re.sub(r"\s+([.;,])", r"\1", text).strip()


def clean(value):
    """Cell to trimmed string. The workbook writes 'None' for an empty cell."""
    if value is None:
        return ""
    text = str(value).strip()
    if text.lower() in ("none", "n/a", "-", ""):
        return ""
    # Mojibake em-dashes from the AI capture.
    return strip_citations(text.replace("\ufffd", "—"))


def split_semicolons(value):
    return [part.strip() for part in clean(value).split(";") if part.strip()]


def split_commas(value):
    """
    Splits a food list on commas that are not inside brackets.

    "Exceptional Kibble (preferred), Raw Mutton, or Raw Meat" is three items;
    splitting naively would tear the bracketed qualifier off the first.
    """
    text = clean(value)
    if not text:
        return []
    parts, depth, current = [], 0, ""
    for ch in text:
        if ch in "([":
            depth += 1
        elif ch in ")]":
            depth = max(0, depth - 1)
        if ch == "," and depth == 0:
            parts.append(current)
            current = ""
        else:
            current += ch
    parts.append(current)
    out = []
    for part in parts:
        item = part.strip().rstrip(".").strip()
        item = re.sub(r"^(?:or|and)\s+", "", item, flags=re.I).strip()
        if item:
            out.append(item)
    return out


def drops(value, prefix):
    """
    Drop rows carrying only what is known.

    Every other field on DropEntrySchema has a default, and package content is
    parsed through that schema on read, so omitting them costs nothing.
    """
    return [
        {"id": f"{prefix}{i + 1}", "label": label}
        for i, label in enumerate(split_semicolons(value))
    ]


def core_name(name):
    """A creature name with its variant qualifiers removed, for comparison."""
    return re.sub(r"[^a-z]", "", re.sub(QUALIFIER, "", name.lower()))


def variant_parents(creatures):
    """
    child blueprint path -> parent blueprint path, by name.

    The bundled catalog declares variant parents for items only, so creature
    variants have never had one — which is why every Aberrant creature carries
    a full duplicate of its base creature's record.
    """
    by_name = {c["name"]: c for c in creatures}
    pairs = {}
    for creature in creatures:
        name = creature["name"]
        base = None
        for prefix in VARIANT_PREFIXES:
            if name.startswith(prefix) and name[len(prefix):] in by_name:
                base = name[len(prefix):]
                break
        if base is None:
            for suffix in VARIANT_SUFFIXES:
                if name.endswith(suffix) and name[: -len(suffix)] in by_name:
                    base = name[: -len(suffix)]
                    break
        if base and base != name:
            pairs[norm(creature["bpPath"])] = norm(by_name[base]["bpPath"])
    return pairs


def guide_text(row, gcol):
    parts = [clean(row[gcol["Weapons"]]), clean(row[gcol["Taming Items"]])]
    for n in range(1, 5):
        parts.append(clean(row[gcol[f"Step {n} Title"]]))
        parts.append(clean(row[gcol[f"Step {n} Method"]]))
    return " ".join(parts)


def distinctive(name):
    """The words a guide about this creature has to actually use."""
    words = re.findall(r"[A-Za-z]+", name)
    keep = [w for w in words if w.lower() not in QUALIFIER and len(w) > 2]
    keep = [w for w in keep if not re.fullmatch(QUALIFIER, w.lower())]
    return [w.lower() for w in (keep or words)]


def names_itself(name, text):
    """
    Whether a guide uses the creature's own name, allowing one typo.

    The capture misspells a few — Cymathoa's guide says "Cymothoa" throughout —
    and a strict match would read that as a guide about somebody else and throw
    away a perfectly good one.
    """
    for token in distinctive(name):
        if re.search(rf"\b{re.escape(token)}", text):
            return True
        if len(token) < 6:
            continue
        # The same word with any single character substituted.
        for i in range(len(token)):
            near = f"{re.escape(token[:i])}.{re.escape(token[i + 1:])}"
            if re.search(rf"\b{near}", text):
                return True
    return False


def misassociated(rows, gcol, names_by_path):
    """
    Rows whose guide is about some other creature.

    Two independent signals, because neither catches everything:

      A guide that is a verbatim copy of another row's, where the two
      creatures are not variants of one another. The copy is dropped from
      every member except the one whose name the text actually uses.

      A guide that never names the creature it is filed under. This catches
      the cases where the text was written about a different creature outright
      rather than copied from a neighbouring row.
    """
    dropped = {}

    groups = defaultdict(list)
    for row in rows:
        markdown = str(row[gcol["Full Markdown"]] or "")
        if markdown:
            groups[markdown].append(row)

    for members in groups.values():
        if len(members) < 2:
            continue
        cores = {core_name(str(m[gcol["Creature Name"]])) for m in members}
        if len(cores) == 1:
            continue  # Variants of one creature legitimately share a guide.
        text = guide_text(members[0], gcol).lower()
        for row in members:
            name = str(row[gcol["Creature Name"]])
            if not names_itself(name, text):
                dropped[norm(clean(row[gcol["Blueprint Path"]]))] = (
                    f"shares {members[0][gcol['Creature Name']]}'s guide"
                )

    # Every distinctive creature word in the catalog, for spotting a guide
    # that is about somebody else.
    owners = defaultdict(set)
    for creature_name in names_by_path.values():
        for token in distinctive(creature_name):
            if len(token) >= 5:
                owners[token].add(core_name(creature_name))

    for row in rows:
        path = norm(clean(row[gcol["Blueprint Path"]]))
        if path in dropped or path not in names_by_path:
            continue
        text = guide_text(row, gcol).lower()
        if not text.strip():
            continue
        name = names_by_path[path]
        if names_itself(name, text):
            continue
        mine = core_name(name)
        # Named twice or more: one passing mention of another creature is
        # ordinary in a taming guide ("clear out nearby Ravagers first").
        elsewhere = sorted(
            token
            for token, cores in owners.items()
            if mine not in cores
            and len(re.findall(rf"\b{re.escape(token)}\b", text)) > 1
        )
        if elsewhere:
            dropped[path] = f"guide is about {elsewhere[0]}"

    return dropped


def same_section(a, b, section):
    """
    Whether two records say the same thing about one section.

    Acquisition compares availability alone, deliberately. Two guides for the
    same creature almost never match word for word — the capture was made per
    row — and treating a reworded copy as a real difference would leave every
    variant carrying its own near-duplicate. What actually differs between a
    Rex and a Ghost Rex is whether it can be obtained at all, so that is the
    test: same answer, same process, inherit it.
    """
    if section == "acquisition":
        return a.get("availability") == b.get("availability")
    if section == "spawns":
        return a.get("spawnMaps", []) == b.get("spawnMaps", [])
    if section == "drops":
        return a.get("drops", {}) == b.get("drops", {})
    if section == "technical":
        return a.get("technical", {}) == b.get("technical", {})
    return True


def main():
    book = openpyxl.load_workbook(WORKBOOK, read_only=True, data_only=True)
    details = list(book["Creature Details"].iter_rows(values_only=True))
    dcol = {name: i for i, name in enumerate(details[0])}
    guides = list(book["Creature Guides"].iter_rows(values_only=True))
    gcol = {name: i for i, name in enumerate(guides[0])}

    catalog = json.load(open(CATALOG, encoding="utf8"))
    names_by_path = {norm(c["bpPath"]): c["name"] for c in catalog["creatures"]}
    parents = variant_parents(catalog["creatures"])

    # Keyed by blueprint path: creature names repeat, paths do not.
    guide_rows = {}
    for row in guides[1:]:
        guide_rows.setdefault(norm(clean(row[gcol["Blueprint Path"]])), row)

    dropped = misassociated(guides[1:], gcol, names_by_path)

    records = {}
    stats = Counter()

    for row in details[1:]:
        path = norm(clean(row[dcol["Blueprint Path"]]))
        if path not in names_by_path:
            stats["skipped-unknown-path"] += 1
            continue

        availability = clean(row[dcol["Availability"]]).lower()
        if availability not in ("acquirable", "unavailable"):
            availability = ""

        spawn_maps = split_semicolons(row[dcol["Spawn Maps"]])

        drag = clean(row[dcol["Drag Weight"]])
        try:
            drag_weight = float(drag) if drag else None
            if drag_weight is not None and drag_weight == int(drag_weight):
                drag_weight = int(drag_weight)
        except ValueError:
            drag_weight = None
            stats["drag-weight-unparsed"] += 1

        record_drops = {
            "harvest": drops(row[dcol["Harvest Drops"]], "dh"),
            "guaranteed": drops(row[dcol["Guaranteed Drops"]], "dg"),
            "random": drops(row[dcol["Random Drops"]], "dr"),
            "production": drops(row[dcol["Production Drops"]], "dp"),
        }
        record_drops = {k: v for k, v in record_drops.items() if v}

        methods = []
        guide = guide_rows.get(path)
        if path in dropped:
            guide = None
            stats["guide-dropped"] += 1
        # A method is only attached to a creature the wiki says can be
        # obtained. The guides sheet has prose for untameable creatures too,
        # and publishing it would assert a taming route that does not exist.
        if guide and availability == "acquirable":
            weapons = clean(guide[gcol["Weapons"]])
            items = clean(guide[gcol["Taming Items"]])
            requirements = "; ".join(part for part in (weapons, items) if part)

            # `referenceType` defaults to text and `role` to taming-food.
            inputs = [
                {"id": f"m1i{i + 1}", "label": label}
                for i, label in enumerate(split_commas(guide[gcol["Taming Food"]]))
            ]

            phases = []
            for n in range(1, 5):
                title = clean(guide[gcol[f"Step {n} Title"]])
                text = clean(guide[gcol[f"Step {n} Method"]])
                if not title and not text:
                    continue
                index = len(phases) + 1
                phase = {"id": f"m1p{index}", "name": title}
                if text:
                    phase["steps"] = [{"id": f"m1p{index}s1", "text": text}]
                phases.append(phase)

            if requirements or inputs or phases:
                method = {"id": "m1", "name": "Taming", "outcome": "direct-tame"}
                if requirements:
                    method["requirements"] = requirements
                if inputs:
                    method["inputs"] = inputs
                if phases:
                    method["phases"] = phases
                methods.append(method)
                stats["methods"] += 1

        record = {"availability": availability}
        if methods:
            record["methods"] = methods
        if spawn_maps:
            record["spawnMaps"] = spawn_maps
        if record_drops:
            record["drops"] = record_drops
        if drag_weight is not None:
            record["technical"] = {"dragWeight": drag_weight}
        records[path] = record

    # Variant reduction, after every record exists so a parent is available.
    for path, record in records.items():
        parent = records.get(parents.get(path, ""))
        filled = []
        if record.get("availability") or record.get("methods"):
            filled.append("acquisition")
        if record.get("spawnMaps"):
            filled.append("spawns")
        if record.get("drops"):
            filled.append("drops")
        if record.get("technical"):
            filled.append("technical")

        if parent is None:
            record["overrides"] = [s for s in SECTIONS if s in filled]
            continue

        # Only what differs from the base creature is kept; the rest is
        # inherited, so a variant stops carrying a near-copy of its parent's
        # taming guide and the two can never drift apart.
        owned = [s for s in filled if not same_section(record, parent, s)]
        for section in SECTIONS:
            if section in filled and section not in owned:
                stats[f"inherited-{section}"] += 1
                if section == "acquisition":
                    record.pop("methods", None)
                    record["availability"] = ""
                elif section == "spawns":
                    record.pop("spawnMaps", None)
                elif section == "drops":
                    record.pop("drops", None)
                elif section == "technical":
                    record.pop("technical", None)
        record["overrides"] = [s for s in SECTIONS if s in owned]
        stats["variants"] += 1

    # A variant that turned out identical to its parent holds nothing worth
    # shipping; the parent's record answers for it.
    empty = [
        path
        for path, record in records.items()
        if not record.get("overrides") and not record.get("availability")
    ]
    for path in empty:
        del records[path]
    stats["variants-fully-inherited"] = len(empty)
    stats["records"] = len(records)

    payload = {
        "source": "official-ark-taming-guides.xlsx",
        "provenance": {
            "availability, spawn maps, drag weight, drops": (
                "ARK Official Community Wiki, revision recorded per row in the "
                "workbook's Field Provenance column"
            ),
            "acquisition methods and steps": (
                "Chrome AI capture (ChatGPT and Gemini), not wiki-verified. "
                "Guides that describe a different creature than the row they "
                "were filed under are dropped, not repaired."
            ),
        },
        "droppedGuides": dict(sorted(dropped.items())),
        "variantParents": dict(sorted(parents.items())),
        "creatureInfo": dict(sorted(records.items())),
    }

    with open(OUT, "w", encoding="utf8", newline="\n") as handle:
        json.dump(payload, handle, indent=2, ensure_ascii=False)
        handle.write("\n")

    print(json.dumps(dict(sorted(stats.items())), indent=2))
    print(f"variant pairs: {len(parents)}   dropped guides: {len(dropped)}")
    print("wrote", OUT)


main()
